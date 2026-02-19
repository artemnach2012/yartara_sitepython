import openpyxl
import re
from pathlib import Path

excel_file = "опт 19.02.26.xlsx"
js_file = "js/script.js"

if not Path(excel_file).exists():
    print(f"Ошибка: файл {excel_file} не найден.")
    exit(1)

wb = openpyxl.load_workbook(excel_file, data_only=True)
sheet = wb['Лист_1']

print("Идёт анализ файла...\n")

# Выведем первые 15 строк, чтобы увидеть структуру
print("Первые 15 строк (первые 15 столбцов):")
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    if i >= 15:
        break
    # Приводим все значения к строке для читаемости
    str_row = [str(cell) if cell is not None else '' for cell in row[:15]]
    print(f"Строка {i+1:2}: " + " | ".join(f"{v:10}" for v in str_row))

print("\n--- Поиск товаров ---\n")

products = []
row_count = 0
for row in sheet.iter_rows(values_only=True):
    row_count += 1
    # Берём предполагаемые колонки: D (3), N (13), O (14) – но можем ошибаться
    name = row[3] if len(row) > 3 else None
    price = row[13] if len(row) > 13 else None
    stock = row[14] if len(row) > 14 else None

    # Для отладки выведем одну из первых товарных строк (например, где есть "Абажур")
    if name and isinstance(name, str) and "Абажур" in name:
        print(f"Найдена строка с 'Абажур':")
        print(f"  name (индекс 3): {name}")
        print(f"  price (индекс 13): {price} (тип {type(price)})")
        print(f"  stock (индекс 14): {stock} (тип {type(stock)})")
        # Также выведем все значения этой строки для проверки
        all_vals = [str(v) if v is not None else '' for v in row]
        print("  Все значения:", all_vals)
        print()

    # Пропускаем строки без названия или служебные
    if not name or not isinstance(name, str):
        continue
    name = name.strip()
    if len(name) < 5:
        continue
    if any(x in name for x in ['Прайс-лист', 'Ценовая группа', 'Артикул', 'Номенклатура']):
        continue

    # Пытаемся преобразовать цену и остаток
    try:
        if price is None or stock is None:
            continue
        # Если это строки, заменяем запятую и убираем пробелы
        if isinstance(price, str):
            price = price.replace(',', '.').replace(' ', '')
        price_val = float(price)
        if isinstance(stock, str):
            stock = stock.replace(',', '').replace(' ', '')
        stock_val = float(stock)
    except (ValueError, TypeError):
        continue

    category = detect_category(name)
    volume = extract_volume(name)
    diameter = extract_diameter(name)

    products.append({
        'name': name,
        'price': round(price_val, 2),
        'stock': stock_val,
        'volume': volume,
        'category': category,
        'diameter': diameter
    })

print(f"Всего обработано строк: {row_count}")
print(f"Найдено товаров: {len(products)}")

if len(products) == 0:
    print("Товары не найдены. Проверьте вывод первых строк – возможно, название, цена или остаток находятся в других столбцах.")
    print("Подсказка: в отладке выше найдите строку с 'Абажур' и посмотрите, в каких колонках лежат цена и остаток.")
    exit(1)

# Генерация JS и замена (как в предыдущем скрипте)
# ... (здесь вставьте оставшуюся часть скрипта – генерацию и замену)

# Генерируем JavaScript-код массива
lines = ["const products = ["]
for p in products:
    safe_name = p['name'].replace('"', '\\"')
    line = f'  {{ name: "{safe_name}", price: {p["price"]}, stock: {p["stock"]}, volume: "{p["volume"]}", category: "{p["category"]}", diameter: "{p["diameter"]}" }},'
    lines.append(line)
lines.append("];")
new_products_block = "\n".join(lines)

# Читаем текущий js-файл
with open(js_file, 'r', encoding='utf-8') as f:
    content = f.read()

# Заменяем старый блок между метками
pattern = r'// PRODUCTS_START.*?// PRODUCTS_END'
replacement = f'// PRODUCTS_START\n{new_products_block}\n// PRODUCTS_END'
new_content = re.sub(pattern, replacement, content, flags=re.DOTALL)

# Записываем обратно
with open(js_file, 'w', encoding='utf-8') as f:
    f.write(new_content)

print(f"Файл {js_file} успешно обновлён.")